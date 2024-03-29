# color_demo.py

from itertools import chain
import openpyxl
from openpyxl.styles import PatternFill

colors = (("00000000", "00FFFFFF", "00FF0000", "0000FF00", "000000FF"),
          ("00FFFF00", "00FF00FF", "0000FFFF", "00000000", "00FFFFFF"),
          ("00FF0000", "0000FF00", "000000FF", "00FFFF00", "00FF00FF"),
          ("0000FFFF", "00800000", "00008000", "00000080", "00808000"),
          )

def color_demo(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active


    for idx,color in enumerate(chain(*colors), 1):
        sheet[f"A{idx}"] =idx
        sheet[f"A{idx}"].fill = PatternFill(start_color=color, end_color=color,
                                       fill_type="solid")


    workbook.save(path)


if __name__ == "__main__":
    color_demo("color_demo.xlsx")
