# using_rules.py

from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle


def applying_rules(path, rule_formula, output_path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active

    yellow = PatternFill(bgColor="00FFFF00")
    diff_style = DifferentialStyle(fill=yellow)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = [rule_formula]

    sheet.conditional_formatting.add("A1:F100", rule)
    workbook.save(output_path)


if __name__ == "__main__":
    applying_rules("ratings.xlsx", rule_formula="IF(AND(NOT(ISBLANK($B1)),$B1<3),TRUE,FALSE)",
                   output_path="rules.xlsx")
